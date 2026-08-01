"""Import real care-home roster spreadsheets (spec 1.4).

The homes plan on Excel, so the first thing Emma has to be able to do is read
what they already wrote. This package parses a workbook into plain dataclasses
(`plan.ParsedRoster`), and `loader.apply` writes one of those into the database -
staff, shift definitions, the roster period and version, every staff × day cell,
the leave those cells record, the task codes, the events row and the floor
assignments.

Layout of the package
---------------------
=================  =========================================================
``vocab``          the duty/leave/rank dictionary, taken from the sheets' own
                   legends - the file to edit when a home adds a code
``cells``          the grammar of one cell (``▲SR A5 + OT x 3 hrs``)
``sheets``         worksheet scanning: find the real grid, not ``max_row``
``home_a``         28-day cycle, two rank sheets, before/after pair
``home_b``         natural month, three floors, paired duty/floor rows
``plan``           the parsed result and its validation summary
``loader``         the only module that talks to the database
=================  =========================================================

Adding a home is one new reader module plus a ``FacilityProfile``; nothing else
in the package or the API changes.
"""
from __future__ import annotations

import hashlib
from typing import Protocol

from . import home_a, home_b
from .plan import Issue, ParsedCell, ParsedEvent, ParsedRoster, ParsedStaff
from .vocab import PROFILES, FacilityProfile


class Layout(Protocol):
    """A layout reader: recognise a workbook, then parse it."""

    LAYOUT: str

    def matches(self, workbook) -> bool: ...
    def read(self, workbook, *, source_name: str) -> ParsedRoster: ...


# Order matters only for ambiguity; the two current layouts are disjoint.
LAYOUTS: tuple = (home_a, home_b)
LAYOUT_NAMES: tuple[str, ...] = tuple(module.LAYOUT for module in LAYOUTS)

__all__ = [
    "FacilityProfile", "Issue", "LAYOUTS", "LAYOUT_NAMES", "ParsedCell",
    "ParsedEvent", "ParsedRoster", "ParsedStaff", "PROFILES", "detect_layout",
    "file_digest", "load_workbook", "parse_workbook",
]


def load_workbook(source):
    """Open a workbook with formulas already evaluated, without modifying it.

    ``source`` is a path, a file-like object, or an already-open workbook, so the
    same call serves the CLI import, the API upload and an in-memory fixture.
    """
    import openpyxl

    if isinstance(source, openpyxl.Workbook):
        return source
    return openpyxl.load_workbook(source, data_only=True, read_only=False)


def detect_layout(workbook):
    """Which reader understands this workbook? None when nothing recognises it."""
    for module in LAYOUTS:
        if module.matches(workbook):
            return module
    return None


def parse_workbook(source, *, source_name: str, variant: str = "after"
                   ) -> ParsedRoster:
    """Parse a roster workbook into a :class:`ParsedRoster`.

    ``variant`` selects Home A's ``before`` or ``after`` sheet pair and is
    ignored by layouts that publish a single grid.
    """
    workbook = load_workbook(source)
    module = detect_layout(workbook)
    if module is None:
        raise ValueError(
            f"{source_name}: unrecognised roster layout (expected one of "
            f"{', '.join(LAYOUT_NAMES)})")
    if module is home_a:
        return module.read(workbook, source_name=source_name, variant=variant)
    return module.read(workbook, source_name=source_name)


def file_digest(path) -> str:
    """SHA-256 of a workbook, so re-importing the same file is visible."""
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 16), b""):
            digest.update(chunk)
    return digest.hexdigest()
