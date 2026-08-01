"""Worksheet scanning helpers shared by the layout readers.

Care-home rosters are laid out for printing, not for parsing: the header row
sits wherever the title block ended, merged cells leave holes, and the sheet's
declared dimensions run to column 4708 because someone once pasted into a far
cell. These helpers locate the real grid instead of trusting `ws.max_row` or a
hard-coded row number, so a home shifting its header down one row does not break
the import.
"""
from __future__ import annotations

import re
from datetime import date as Date

from openpyxl.utils import get_column_letter

# 'RCW 院舍護理員 after!M14' - the reference an import issue points a human at.
def cell_ref(sheet_title: str, row: int, column: int) -> str:
    return f"{sheet_title}!{get_column_letter(column)}{row}"


def text(ws, row: int, column: int) -> str:
    """Trimmed cell text, empty for blanks. Never raises on an empty sheet."""
    value = ws.cell(row=row, column=column).value
    return "" if value is None else re.sub(r"\s+", " ", str(value)).strip()


def used_extent(ws, *, max_scan_rows: int = 400, max_scan_cols: int = 80
                ) -> tuple[int, int]:
    """Last row/column that actually holds a value, ignoring phantom dimensions."""
    last_row = last_col = 0
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, max_scan_rows),
                            max_col=min(ws.max_column, max_scan_cols)):
        for cell in row:
            if cell.value not in (None, ""):
                last_row = max(last_row, cell.row)
                last_col = max(last_col, cell.column)
    return last_row, last_col


# ── date headers ─────────────────────────────────────────────────────────────
_DDMM_RE = re.compile(r"^(\d{1,2})\s*/\s*(\d{1,2})$")
_TITLE_RANGE_RE = re.compile(
    r"(?P<y1>20\d{2})\s*年\s*(?P<m1>\d{1,2})\s*月\s*(?P<d1>\d{1,2})\s*日"
    r".{0,6}?(?:(?P<y2>20\d{2})\s*年\s*)?(?P<m2>\d{1,2})\s*月\s*(?P<d2>\d{1,2})\s*日"
)
_TITLE_MONTH_RE = re.compile(r"(?P<year>20\d{2})\s*年\s*(?P<month>\d{1,2})\s*月")


def find_title_period(ws, *, max_rows: int = 6) -> tuple[Date, Date] | None:
    """Read the printed date range out of the sheet's title block.

    Home A prints '2026年3月2日 至 2026年3月29日'; Home B prints '2026年 6 月',
    which resolves to that whole calendar month.
    """
    import calendar

    for row in range(1, max_rows + 1):
        for column in range(1, 12):
            value = text(ws, row, column)
            if not value:
                continue
            if m := _TITLE_RANGE_RE.search(value):
                year1 = int(m["y1"])
                year2 = int(m["y2"] or year1)
                start = Date(year1, int(m["m1"]), int(m["d1"]))
                end = Date(year2, int(m["m2"]), int(m["d2"]))
                if end >= start:
                    return start, end
            if m := _TITLE_MONTH_RE.search(value):
                year, month = int(m["year"]), int(m["month"])
                last = calendar.monthrange(year, month)[1]
                return Date(year, month, 1), Date(year, month, last)
    return None


def find_date_header(ws, *, year: int, month_hint: int | None = None,
                     max_rows: int = 12, max_cols: int = 60,
                     min_days: int = 20) -> tuple[int, dict[int, Date]] | None:
    """Locate the row of day headers and map each column to a real date.

    Handles both dialects: Home A writes ``02/03``..``29/03`` (dd/mm, so a cycle
    can cross a month boundary), Home B writes bare day numbers ``1``..``30``.
    """
    for row in range(1, max_rows + 1):
        by_column: dict[int, Date] = {}
        day_numbers: dict[int, int] = {}
        for column in range(1, max_cols + 1):
            raw = ws.cell(row=row, column=column).value
            if raw is None:
                continue
            if isinstance(raw, Date):
                by_column[column] = Date(raw.year, raw.month, raw.day)
                continue
            value = re.sub(r"\s+", "", str(raw))
            if m := _DDMM_RE.match(value):
                day, mon = int(m.group(1)), int(m.group(2))
                if 1 <= day <= 31 and 1 <= mon <= 12:
                    by_column[column] = Date(year, mon, day)
                continue
            if value.isdigit() and 1 <= int(value) <= 31:
                day_numbers[column] = int(value)

        if len(by_column) >= min_days:
            return row, _roll_year(by_column)
        # Bare day numbers must be consecutive to be a date header rather than a
        # row of counts, and the month comes from the title block.
        if month_hint and len(day_numbers) >= min_days and _consecutive(day_numbers):
            columns = sorted(day_numbers)
            resolved = {
                column: Date(year, month_hint, day_numbers[column])
                for column in columns
            }
            return row, resolved
    return None


def _consecutive(day_numbers: dict[int, int]) -> bool:
    values = [day_numbers[c] for c in sorted(day_numbers)]
    return values == list(range(values[0], values[0] + len(values)))


def _roll_year(by_column: dict[int, Date]) -> dict[int, Date]:
    """A dd/mm cycle that wraps December→January belongs to the next year."""
    columns = sorted(by_column)
    out: dict[int, Date] = {}
    previous: Date | None = None
    for column in columns:
        current = by_column[column]
        if previous and current < previous:
            current = Date(current.year + 1, current.month, current.day)
        out[column] = current
        previous = current
    return out


# ── row classification ───────────────────────────────────────────────────────
_TIME_LEGEND_RE = re.compile(r"\d\s*[:：]\s*\d{2}\s*(am|pm)?", re.I)
_PAREN_TOKEN_RE = re.compile(r"[（(]\s*([A-Za-z /]{1,12})\s*[)）]")


def looks_like_legend(value: str) -> bool:
    """Group headers repeat the shift times; colour keys explain the fill."""
    return bool(_TIME_LEGEND_RE.search(value)) or "顏色" in value


def looks_like_events_row(value: str) -> bool:
    return any(token in value for token in ("活動", "會議", "備註", "備 註"))


# The homes label their own header rows; those labels are never a staff member.
_HEADER_LABELS = ("姓名", "日期", "職位", "職級", "年份")
_WEEKDAYS = frozenset("一二三四五六日")


def looks_like_header_label(value: str) -> bool:
    return any(token in value for token in _HEADER_LABELS)


def looks_like_weekday_row(values: list[str]) -> bool:
    """The 一二三四五六日 row sits under the dates and holds no duties."""
    filled = [v for v in values if v]
    if len(filled) < 5:
        return False
    return all(v in _WEEKDAYS for v in filled)


def parenthesised_ranks(value: str) -> list[str]:
    """'助理院長 (AS) (RN)' -> ['AS', 'RN'] so a titled row still resolves."""
    return [m.strip() for m in _PAREN_TOKEN_RE.findall(value) if m.strip()]


def is_personal_name(value: str) -> bool:
    """A Chinese personal name (2-4 characters, no code punctuation or digits)."""
    if not (2 <= len(value) <= 4):
        return False
    return all("一" <= ch <= "鿿" for ch in value)


def to_number(value: str) -> float | None:
    try:
        return float(re.sub(r"[^\d.\-]", "", value) or "")
    except ValueError:
        return None
