"""Excel and PDF rendering for reports (spec 3.3 / 7.1 / 7.2).

Every generator in `reports.py` returns the same `{meta, columns, rows}` shape,
which is what made one CSV renderer enough. The same property makes one XLSX
renderer and one PDF renderer enough - nothing here knows what a roster or a
compliance summary is, only how to lay out a table.

Why both formats, and why they are not the same document
--------------------------------------------------------
They are read by different people for different reasons, so they are built
differently on purpose:

  XLSX  a manager opens it and works on it. Frozen header, filters, real numbers
        in number cells - if hours arrive as text nobody can sum a column, which
        is the first thing anyone does with an hours report.

  PDF   goes to SWD, or into a file, or gets signed. It has to be fixed, paginated,
        and self-describing a year later - which period, which roster version,
        generated when, by whom.

The draft watermark
-------------------
A PDF of a *draft* roster is the dangerous artefact in this module. It looks
exactly as official as a published one, and once it is printed or emailed nobody
can tell. So an unpublished roster is watermarked across every page and banded in
the header, and the same status is written into the XLSX header block. It cannot
be exported *silently* as a draft; it can be exported, which matters because
schedulers circulate drafts for comment all the time.

Traditional Chinese
-------------------
Reportlab's built-in `MSung-Light` CID font covers it with no font file to ship.
Without it every Chinese label in a PDF renders as a black box - and every label
on this system can be Chinese.
"""
from __future__ import annotations

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.cidfonts import UnicodeCIDFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

# ── fonts ────────────────────────────────────────────────────────────────────
# `MSung-Light` is a CID font shipped inside reportlab, so there is no .ttf to
# vendor, no licence to track, and no difference between a developer laptop and
# the container.
CJK_FONT = "MSung-Light"

# The CMap that maps Unicode to CIDs in the Adobe-**CNS1** collection, which is
# the Traditional Chinese one and the collection MSung-Light belongs to.
#
# We set this explicitly because reportlab's own default is wrong for this face.
# `_cidfontdata.defaultUnicodeEncodings` maps MSung-Light to
# ('cht', 'UniGB-UCS2-H') - correctly labelled Traditional, then paired with the
# Adobe-**GB1** CMap, which is Simplified. The CID numbers that produces index
# into the wrong character collection, so the glyphs that come out are not the
# characters that went in.
#
# It renders. It just renders the wrong Chinese, which is the worst kind of
# encoding bug: nothing raises, the PDF opens, and only a Chinese reader notices.
CJK_ENCODING = "UniCNS-UCS2-H"
_FONTS_READY = False


def _ensure_fonts() -> str:
    """Register the CJK font once, with the Traditional Chinese CMap."""
    global _FONTS_READY
    if not _FONTS_READY:
        from reportlab.pdfbase import _cidfontdata

        language, _wrong = _cidfontdata.defaultUnicodeEncodings[CJK_FONT]
        _cidfontdata.defaultUnicodeEncodings[CJK_FONT] = (language, CJK_ENCODING)
        pdfmetrics.registerFont(UnicodeCIDFont(CJK_FONT))
        _FONTS_READY = True
    return CJK_FONT


# ── shared helpers ───────────────────────────────────────────────────────────
BRAND = "2C3E50"
HEADER_FILL = "34495E"
BAND_FILL = "F4F6F8"
DRAFT_RED = "C0392B"

# Rendering is capped so one bad parameter cannot try to lay out a 200 000-row
# PDF and take the API process with it. The cap is reported in the document
# rather than applied quietly - a truncated report that does not say so is worse
# than no report.
MAX_PDF_ROWS = 2000


def _status(payload: dict) -> str:
    return str((payload.get("meta") or {}).get("roster_version_status") or "").lower()


def is_draft(payload: dict) -> bool:
    """True when this report describes a roster that has not been published.

    Reports with no roster version at all (a staff register, an evidence pack)
    are not drafts - there is no roster behind them to be unpublished.
    """
    status = _status(payload)
    return bool(status) and status != "published"


def _cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def _numeric(value):
    """Return a real number when the value is one, else None.

    Booleans are excluded deliberately: `True` is an int in Python, and an hours
    column with a stray `1` where a Yes belongs is the kind of error that only
    shows up once someone sums the column.
    """
    if isinstance(value, bool) or value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    try:
        return float(text) if "." in text else int(text)
    except ValueError:
        return None


def _meta_lines(payload: dict, *, title: str) -> list[tuple[str, str]]:
    meta = payload.get("meta") or {}
    lines: list[tuple[str, str]] = [("Report", title)]
    if meta.get("description"):
        lines.append(("Description", str(meta["description"])))
    start = meta.get("period_start") or meta.get("date_from")
    end = meta.get("period_end") or meta.get("date_to")
    if start or end:
        lines.append(("Period", f"{start or '?'} to {end or '?'}"))
    if meta.get("roster_version_label"):
        lines.append(("Roster version", str(meta["roster_version_label"])))
    if _status(payload):
        lines.append(("Roster status", _status(payload).upper()))
    lines.append(("Rows", str(len(payload.get("rows") or ()))))
    lines.append(("Generated", datetime.now(timezone.utc)
                  .strftime("%Y-%m-%d %H:%M UTC")))
    return lines


# ── XLSX ─────────────────────────────────────────────────────────────────────
def to_xlsx(payload: dict, *, title: str = "Report") -> bytes:
    """One sheet: a meta block, then the table with a frozen, filterable header.

    Numbers are written as numbers. It sounds trivial and it is the whole point of
    shipping XLSX rather than renaming a CSV - a column of text-formatted hours
    cannot be summed, sorted or charted, and summing hours is the first thing any
    manager does with this file.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = (title[:28] or "Report").replace("/", "-").replace("\\", "-")

    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    width = max(len(columns), 2)

    bold = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=HEADER_FILL)
    thin = Side(style="thin", color="D5DBDB")
    border = Border(bottom=thin)

    line = 1
    for label, value in _meta_lines(payload, title=title):
        sheet.cell(row=line, column=1, value=label).font = Font(bold=True, color=BRAND)
        cell = sheet.cell(row=line, column=2, value=value)
        if label == "Roster status" and is_draft(payload):
            cell.font = Font(bold=True, color=DRAFT_RED)
        line += 1

    if is_draft(payload):
        warn = sheet.cell(
            row=line, column=1,
            value="DRAFT - this roster has not been published. "
                  "Do not treat as the operative roster.")
        warn.font = Font(bold=True, color=DRAFT_RED)
        sheet.merge_cells(start_row=line, start_column=1,
                          end_row=line, end_column=width)
        line += 1
    line += 1

    header_row = line
    for index, column in enumerate(columns, start=1):
        cell = sheet.cell(row=header_row, column=index,
                          value=column.get("label") or column.get("key"))
        cell.font, cell.fill, cell.border = bold, head_fill, border
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)

    keys = [c["key"] for c in columns]
    widths = [len(str(c.get("label") or c["key"])) for c in columns]
    for offset, row in enumerate(rows, start=1):
        for index, key in enumerate(keys, start=1):
            raw = row.get(key)
            number = _numeric(raw)
            cell = sheet.cell(row=header_row + offset, column=index,
                              value=number if number is not None else _cell(raw))
            if number is not None:
                cell.alignment = Alignment(horizontal="right")
            widths[index - 1] = max(widths[index - 1], len(_cell(raw)))

    for index, size in enumerate(widths, start=1):
        # Bounded: an unbounded width makes a single long note push every other
        # column off the screen.
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(size + 2, 9), 42)

    if columns:
        sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)
        if rows:
            sheet.auto_filter.ref = (
                f"A{header_row}:"
                f"{get_column_letter(len(columns))}{header_row + len(rows)}")

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


# ── PDF ──────────────────────────────────────────────────────────────────────
def _page_furniture(payload: dict, title: str):
    """Draw the watermark, the rule and the footer on every page.

    A callback rather than a flowable because it has to fire for pages the story
    creates on its own - a table that spills onto page four still needs the DRAFT
    watermark on page four.
    """
    draft = is_draft(payload)
    font = _ensure_fonts()

    def draw(canvas, doc):
        canvas.saveState()
        page_width, page_height = doc.pagesize

        if draft:
            canvas.setFont(font, 62)
            canvas.setFillColor(colors.HexColor("#" + DRAFT_RED), alpha=0.10)
            canvas.translate(page_width / 2, page_height / 2)
            canvas.rotate(38)
            canvas.drawCentredString(0, 0, "DRAFT")
            canvas.rotate(-38)
            canvas.translate(-page_width / 2, -page_height / 2)

        canvas.setFillColor(colors.HexColor("#7F8C8D"))
        canvas.setFont(font, 7.5)
        canvas.drawString(15 * mm, 10 * mm, title)
        canvas.drawRightString(page_width - 15 * mm, 10 * mm,
                               f"Page {doc.page}")
        canvas.setStrokeColor(colors.HexColor("#D5DBDB"))
        canvas.line(15 * mm, 13 * mm, page_width - 15 * mm, 13 * mm)
        canvas.restoreState()

    return draw


def to_pdf(payload: dict, *, title: str = "Report") -> bytes:
    """A fixed, paginated, self-describing version of the same table.

    Orientation follows the shape of the data: a roster export is staff by days
    and unreadable in portrait, a compliance summary is four columns and wasteful
    in landscape. Choosing per report rather than per format is the difference
    between a PDF someone reads and one they ask to be re-sent as a spreadsheet.
    """
    font = _ensure_fonts()
    columns = payload.get("columns") or []
    rows = list(payload.get("rows") or [])
    truncated = max(0, len(rows) - MAX_PDF_ROWS)
    if truncated:
        rows = rows[:MAX_PDF_ROWS]

    wide = len(columns) > 8
    pagesize = landscape(A4) if wide else A4
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=pagesize,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=14 * mm, bottomMargin=16 * mm,
        title=title, author="Emma AI",
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Title"], fontName=font,
                        fontSize=15, spaceAfter=2, textColor=colors.HexColor("#" + BRAND))
    small = ParagraphStyle("small", parent=styles["Normal"], fontName=font,
                           fontSize=8, textColor=colors.HexColor("#5D6D7E"))
    warn = ParagraphStyle("warn", parent=styles["Normal"], fontName=font,
                          fontSize=9.5, alignment=TA_CENTER,
                          textColor=colors.white, backColor=colors.HexColor("#" + DRAFT_RED),
                          borderPadding=4, spaceBefore=4, spaceAfter=6)
    cell_style = ParagraphStyle("cell", parent=styles["Normal"], fontName=font,
                                fontSize=7 if wide else 8, leading=9)
    head_style = ParagraphStyle("head", parent=cell_style, textColor=colors.white,
                                fontSize=7 if wide else 8)

    story: list = [Paragraph(title, h1)]
    for label, value in _meta_lines(payload, title=title):
        if label == "Report":
            continue
        story.append(Paragraph(f"<b>{label}:</b> {value}", small))
    if is_draft(payload):
        story.append(Paragraph(
            "DRAFT — this roster has not been published. "
            "Do not treat as the operative roster.", warn))
    if truncated:
        story.append(Paragraph(
            f"<b>Showing the first {MAX_PDF_ROWS:,} of "
            f"{MAX_PDF_ROWS + truncated:,} rows.</b> Export as XLSX or CSV for "
            "the complete set.", warn))
    story.append(Spacer(1, 5))

    if not columns:
        story.append(Paragraph("This report produced no columns.", small))
    else:
        header = [Paragraph(f"<b>{c.get('label') or c['key']}</b>", head_style)
                  for c in columns]
        keys = [c["key"] for c in columns]
        body = [[Paragraph(_cell(row.get(k)), cell_style) for k in keys]
                for row in rows]
        if not body:
            body = [[Paragraph("No rows.", cell_style)] + [""] * (len(keys) - 1)]

        table = Table([header, *body], repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#" + HEADER_FILL)),
            ("VALIGN",       (0, 0), (-1, -1), "TOP"),
            ("GRID",         (0, 0), (-1, -1), 0.25, colors.HexColor("#D5DBDB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#" + BAND_FILL)]),
            ("LEFTPADDING",  (0, 0), (-1, -1), 3),
            ("RIGHTPADDING", (0, 0), (-1, -1), 3),
            ("TOPPADDING",   (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        story.append(table)

    furniture = _page_furniture(payload, title)
    doc.build(story, onFirstPage=furniture, onLaterPages=furniture)
    return buffer.getvalue()


# ── format registry ──────────────────────────────────────────────────────────
# The router reads this rather than branching on the extension, so adding a
# format is one entry and no endpoint changes.
FORMATS: dict[str, dict] = {
    "csv": {
        "media_type": "text/csv; charset=utf-8",
        "binary": False,
    },
    "xlsx": {
        "media_type": ("application/vnd.openxmlformats-officedocument."
                       "spreadsheetml.sheet"),
        "binary": True,
        "render": to_xlsx,
    },
    "pdf": {
        "media_type": "application/pdf",
        "binary": True,
        "render": to_pdf,
    },
}
