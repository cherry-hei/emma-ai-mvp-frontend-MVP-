"""Roster import tests (spec 1.4).

Three layers, so a failure says where the problem is:

* **Cell grammar** - pure, no database, no files. The table of cases is taken
  verbatim from the source workbooks, including the typing slips in them.
* **Layout readers** - against workbooks built in memory, so the readers are
  tested without shipping a care home's real roster into the repository.
* **Real workbooks** - run only when `docs/` is present (it is gitignored, being
  real staff data). These assert the two homes' actual files parse with no
  unresolved cells, which is the acceptance criterion for 1.4.
"""
from __future__ import annotations

import pathlib

import pytest

from emma_core.importers import PROFILES, detect_layout, load_workbook, parse_workbook
from emma_core.importers.cells import parse_cell, parse_subrow
from emma_core.importers.vocab import HOME_A, HOME_B, resolve_rank

HOME_A_NURSING = HOME_A.sheet("nursing")
HOME_A_CARE = HOME_A.sheet("care")
HOME_B_SHEET = HOME_B.sheets[0]

DOCS = pathlib.Path(__file__).resolve().parent.parent.parent / "docs"


# ── cell grammar ─────────────────────────────────────────────────────────────
@pytest.mark.parametrize("raw,shift,task", [
    ("A", "A", None),
    ("P", "P", None),
    ("A7", "A", "A7"),                    # task code written into the duty
    ("P4", "P", "P4"),
    ("N3", "N", "N3"),
    ("A/N", "AN", None),
    ("A2N", "AN", "A2"),                  # A/N whose morning half is task A2
    ("#A3N*", "AN", "A3"),                # event markers strip away
    ("長A7", "LA", "A7"),                 # Home A's long duty
    ("E2", "E", None),                    # E/E2/E3 share one window
])
def test_home_a_duty_codes(raw, shift, task):
    intent = parse_cell(raw, HOME_A, HOME_A_CARE)
    assert [d.shift_code for d in intent.duties] == [shift]
    assert intent.duties[0].task_code == task
    assert intent.is_working


@pytest.mark.parametrize("raw,shift", [
    ("7A", "7A"), ("7P", "7P"), ("9A", "9A"), ("D", "D"),
    ("A#", "A"), ("P*", "P"), ("A#/N#", "AN"), ("Run A", "A"), ("Run P", "P"),
])
def test_home_b_duty_codes(raw, shift):
    intent = parse_cell(raw, HOME_B, HOME_B_SHEET)
    assert [d.shift_code for d in intent.duties] == [shift]


@pytest.mark.parametrize("raw,code,hours", [
    ("休", "OFF", None), ("年", "AL", None), ("大", "VL", None),
    ("法", "SH", None), ("公", "PH", None), ("病", "SL", None),
    ("補休", "CL", None), ("補DO", "CO", None), ("S", "SLEEP", None),
    ("ND", "ND", None), ("民", "CIVIC", None), ("救", "TRN", None),
    ("DSL", "DSL", None), ("FAL", "FAL", None),
    ("SH9", "SH", 9.0), ("PH11", "PH", 11.0), ("CL8.5", "CL", 8.5),
])
def test_leave_codes(raw, code, hours):
    intent = parse_cell(raw, HOME_A, HOME_A_NURSING)
    assert intent.leave and intent.leave.code == code
    assert intent.leave_hours == hours
    assert not intent.is_working


def test_rest_day_counters_are_sequence_numbers_not_hours():
    """Home B numbers its rest days DO1..DO7 through the cycle."""
    intent = parse_cell("DO3", HOME_B, HOME_B_SHEET)
    assert intent.leave.code == "DO"
    assert intent.leave_sequence == 3 and intent.leave_hours is None


def test_staff_request_prefix_is_recognised_without_a_separator():
    """The homes write SR against a cell the staff member asked for - and they
    write it glued to the code, where a word boundary does not help."""
    for raw in ("SR休", "SR 休", "SR補休", "SR民 / +CL4", "SR A8"):
        assert parse_cell(raw, HOME_A, HOME_A_CARE).is_request, raw
    # 'SRN' is Home B's senior-nurse rank, not a request marker.
    assert resolve_rank("SRN").rank == "RN"


def test_overtime_and_compensatory_hours():
    intent = parse_cell("A + OT x 4 hrs", HOME_A, HOME_A_CARE)
    assert intent.duties[0].shift_code == "A" and intent.ot_minutes == 240
    assert parse_cell("P1 + CL x 0.5 hr", HOME_A, HOME_A_CARE).cl_minutes == 30
    assert parse_cell("CL 20 mins", HOME_A, HOME_A_CARE).leave.code == "CL"
    # A cell that is only compensatory hours is a compensatory day.
    assert parse_cell("CL8", HOME_A, HOME_A_CARE).leave.code == "CL"


def test_typing_slips_in_the_source_are_repaired():
    """'=' and '_' appear where every sibling cell writes '+', and 'o.5' is a
    letter o typed for a zero."""
    assert parse_cell("A1 = OT x 4 + P2", HOME_A, HOME_A_CARE).ot_minutes == 240
    assert parse_cell("P2 _ CL x 1 hr", HOME_A, HOME_A_CARE).cl_minutes == 60
    assert parse_cell("08:00-16:00 + OT x o.5 hrs",
                      HOME_A, HOME_A_CARE).ot_minutes == 30


def test_double_duty_becomes_one_shift_with_two_segments():
    """'A3 + P3' is two disjoint windows in a day - the A/N shape, not two cells."""
    intent = parse_cell("A3 OT / P2", HOME_A, HOME_A_CARE)
    assert len(intent.duties) == 1
    duty = intent.duties[0]
    assert duty.shift_code == "AP"
    assert [(s["start"], s["end"]) for s in duty.segments] == [
        ("07:00", "15:00"), ("13:30", "21:30")]


@pytest.mark.parametrize("raw,window", [
    ("09:30-19:00", ("09:30", "19:00")),
    ("0830-1630", ("08:30", "16:30")),
    ("8-4", ("08:00", "16:00")),          # a smaller end hour means the afternoon
    ("7-11", ("07:00", "11:00")),
    ("9A-1P", ("09:00", "13:00")),
    ("1-5 PM", ("13:00", "17:00")),
])
def test_explicit_clock_times(raw, window):
    intent = parse_cell(raw, HOME_A, HOME_A_CARE)
    assert intent.time_windows == (window,)
    # Hours the home wrote outside its own dictionary are still a duty.
    assert intent.duties[0].shift_code == "AH"
    assert (intent.duties[0].start, intent.duties[0].end) == window


def test_part_day_leave_keeps_both_halves():
    """'DSL AM / 1-5 PM' - compassionate sick leave in the morning, worked after."""
    intent = parse_cell("DSL AM / 1-5 PM", HOME_A, HOME_A_CARE)
    assert intent.leave.code == "DSL"
    assert intent.day_part == "AM"
    assert intent.time_windows == (("13:00", "17:00"),)


def test_mixed_script_cell_splits_on_the_script_boundary():
    """'B節4' is the B duty plus four hours of festival leave."""
    intent = parse_cell("B節4", HOME_A, HOME_A_CARE)
    assert intent.duties[0].shift_code == "B"
    assert intent.leave.code == "FL" and intent.leave_hours == 4.0


def test_floor_written_into_a_duty_cell_is_a_unit_hint():
    assert parse_cell("2/F P", HOME_B, HOME_B_SHEET).unit_hint == "2/F"
    assert parse_cell("6/F P", HOME_B, HOME_B_SHEET).duties[0].shift_code == "P"
    # '1,6/F A' is two floors, not a decimal.
    assert parse_cell("1,6/F A", HOME_B, HOME_B_SHEET).unit_hint == "1/F"


def test_unresolvable_fragment_is_reported_not_dropped():
    intent = parse_cell("ZZ9", HOME_A, HOME_A_CARE)
    assert intent.unparsed and not intent.duties and not intent.leave


def test_subrow_carries_floor_and_standing_duty():
    note = parse_subrow("6/F+C")
    assert note.floor == "6/F" and note.tasks == ("Canteen",)
    assert parse_subrow("R").tasks == ("Runner",)
    assert parse_subrow("AA").tasks == ("Assist exercise",)


@pytest.mark.parametrize("label,rank,employment", [
    ("RN1", "RN", "local_ft"),
    ("EN2", "EN", "local_ft"),
    ("HW9", "HW", "local_ft"),
    ("RCW12", "CW", "local_ft"),
    ("PT RCW1", "CW", "local_pt"),
    ("HW I", "HW", "imported_labor"),
    ("HCA I", "HCA", "imported_labor"),
    ("外判", "HCA", "outsource"),
    ("替假", "HCA", "casual"),
])
def test_rank_labels_resolve_to_schema_values(label, rank, employment):
    spec = resolve_rank(label)
    assert spec and (spec.rank, spec.employment_type) == (rank, employment)


def test_profiles_expose_one_dictionary_per_facility():
    for code, profile in PROFILES.items():
        codes = [w.code for w in profile.shift_windows]
        assert len(codes) == len(set(codes)), f"Home {code} has a duplicate duty code"
        assert "AN" in codes and profile.double_duty.code in codes


def test_home_a_and_home_b_an_shifts_match_the_specification():
    """Home A A/N = 6.5h + 9.5h = 16h; Home B = 7.5h + 10h = 17.5h."""
    from emma_core.shifttime import paid_minutes

    a = HOME_A_NURSING.window("AN")
    b = HOME_B_SHEET.window("AN")
    assert paid_minutes({"segments": [dict(s) for s in a.segments]}) == 960
    assert paid_minutes({"segments": [dict(s) for s in b.segments]}) == 1050


# ── layout readers, against an in-memory workbook ────────────────────────────
def _home_a_workbook():
    """A minimal Home A sheet: title block, day header, one staff row, events."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RCW 院舍護理員 after"
    ws["A1"] = "2026年3月2日 至 2026年3月29日 員工工作時間表 — RCW 院舍護理員"
    ws["A4"] = "日期/年份 2026"
    for index in range(28):
        column = 2 + index
        day = 2 + index
        ws.cell(4, column).value = f"{day:02d}/03" if day <= 29 else None
    ws["A5"] = "姓名及職位"
    ws["A7"] = "RCW1"
    codes = ["A7", "P4", "A2N", "S", "補休", "休", "年"] * 4
    for index, code in enumerate(codes):
        ws.cell(7, 2 + index).value = code
    ws["A9"] = "活動／會議"
    ws.cell(9, 2).value = "▲全體職員會議"
    return wb


def test_home_a_reader_parses_period_staff_cells_and_events():
    parsed = parse_workbook(_home_a_workbook(), source_name="fixture.xlsx")
    assert parsed.facility_code == "A"
    assert (parsed.period_start.isoformat(), parsed.period_end.isoformat()) == (
        "2026-03-02", "2026-03-29")
    assert len(parsed.dates) == 28
    assert [s.display_name for s in parsed.staff] == ["RCW1"]
    assert parsed.staff[0].rank == "CW"
    assert len(parsed.cells) == 28
    assert parsed.events and parsed.events[0].markers == ("▲",)
    assert not [i for i in parsed.issues if i.severity != "info"]

    summary = parsed.summary()
    assert summary["working_cells"] == 12          # A7, P4, A2N x 4 cycles
    assert summary["leave_cells"] == 16            # S, 補休, 休, 年 x 4 cycles
    assert summary["shift_type_counts"] == {"A": 4, "AN": 4, "P": 4}


def _home_b_workbook():
    """A minimal Home B sheet: title, day numbers, a duty row and its floor row."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "2026"
    ws["D1"] = "迎進生活 2026年 6 月 1/F, 2/F, 6/F 護士/保健員更期表"
    ws["B3"] = "職級"
    ws["C3"] = "姓名"
    for index in range(30):
        ws.cell(3, 4 + index).value = index + 1
    ws["B6"] = "HCA"
    ws["C6"] = "陳測試"
    for index in range(30):
        ws.cell(6, 4 + index).value = "7A" if index % 5 else "DO"
    ws["B7"] = "#"
    ws["C7"] = "6/F"
    ws.cell(7, 4).value = "6/F+C"
    ws["C34"] = "Quoto for request O"
    for index in range(30):
        ws.cell(34, 4 + index).value = 2
    return wb


def test_home_b_reader_parses_floors_standing_duties_and_quota():
    parsed = parse_workbook(_home_b_workbook(), source_name="fixture.xlsx")
    assert parsed.facility_code == "B"
    assert (parsed.period_start.isoformat(), parsed.period_end.isoformat()) == (
        "2026-06-01", "2026-06-30")
    assert [s.display_name for s in parsed.staff] == ["陳測試"]
    assert len(parsed.cells) == 30

    first = next(c for c in parsed.cells if c.date.day == 1)
    assert first.unit_name == "6/F"
    assert first.extra_tasks == ("Canteen",)
    assert len(parsed.request_quota) == 30


def test_unrecognised_workbook_is_rejected_with_a_useful_message():
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active["A1"] = "shopping list"
    with pytest.raises(ValueError, match="unrecognised roster layout"):
        parse_workbook(wb, source_name="notaroster.xlsx")


def test_layout_detection_is_unambiguous():
    assert detect_layout(_home_a_workbook()).LAYOUT == "home_a_duty_roster"
    assert detect_layout(_home_b_workbook()).LAYOUT == "home_b_floor_roster"


# ── the homes' real workbooks (skipped when docs/ is absent) ──────────────────
REAL_SOURCES = [
    ("Duty_Roster_March2026.xlsx", "after", "A", "2026-03-02", "2026-03-29"),
    ("Duty_Roster_March2026.xlsx", "before", "A", "2026-03-02", "2026-03-29"),
    ("FL Nursing Staff Roster 062026.xlsx", "after", "B", "2026-06-01", "2026-06-30"),
    ("night roster.xlsx", "after", "B", "2026-07-01", "2026-07-31"),
]


@pytest.mark.parametrize("name,variant,facility,start,end", REAL_SOURCES)
def test_real_workbooks_parse_without_unresolved_cells(name, variant, facility,
                                                       start, end):
    """Spec 1.4 acceptance: the real sample rosters import with a clean summary.

    `docs/` holds real staff data and is gitignored, so this skips where the files
    are not present rather than shipping them.
    """
    path = DOCS / name
    if not path.exists():
        pytest.skip(f"{name} not present in {DOCS}")
    parsed = parse_workbook(path, source_name=name, variant=variant)

    assert parsed.facility_code == facility
    assert (parsed.period_start.isoformat(), parsed.period_end.isoformat()) == (start, end)
    assert parsed.staff and parsed.cells
    blocking = [i for i in parsed.issues if i.severity in ("warning", "error")]
    assert not blocking, "unresolved cells: " + "; ".join(
        f"{i.cell_ref}={i.raw_value!r}" for i in blocking[:10])

    # Every parsed cell resolves to something writable: a duty or a leave code.
    undecided = [c for c in parsed.cells
                 if not c.intent.duties and not c.intent.leave and not c.relief_name]
    assert not undecided, f"{len(undecided)} cell(s) carried neither duty nor leave"


def test_real_home_a_before_and_after_differ():
    """Home A publishes the cycle twice; the pair is what makes the import useful.

    The as-worked sheet carries the overtime the plan did not, which is the whole
    reason the home keeps both.
    """
    path = DOCS / "Duty_Roster_March2026.xlsx"
    if not path.exists():
        pytest.skip(f"Duty_Roster_March2026.xlsx not present in {DOCS}")
    plan = parse_workbook(path, source_name="plan", variant="before").summary()
    worked = parse_workbook(path, source_name="worked", variant="after").summary()

    assert plan["staff_rows"] == worked["staff_rows"]
    assert worked["overtime_minutes"] > plan["overtime_minutes"]


def test_real_workbook_opens_read_only_without_mutating_it():
    path = DOCS / "Duty_Roster_March2026.xlsx"
    if not path.exists():
        pytest.skip(f"Duty_Roster_March2026.xlsx not present in {DOCS}")
    before = path.stat().st_mtime
    load_workbook(path)
    assert path.stat().st_mtime == before
