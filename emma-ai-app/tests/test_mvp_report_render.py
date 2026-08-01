"""Excel and PDF report rendering - MVP 3.3 / 7.1 / 7.2.

Task 3.3 asks for "export basic Excel/PDF report". The signed scope said data
only, JSON and CSV; Kien's call on 1 Aug was to build the rendering layer rather
than let the demo fail on a disagreement about scope.

What these tests defend is the handful of things that make a rendered report
usable rather than merely produced:

  * numbers arrive in Excel as numbers, because summing the hours column is the
    first thing anyone does with an hours report;
  * Traditional Chinese survives into the PDF, because every label on this system
    can be Chinese and the failure mode is a page of black boxes;
  * a draft roster is marked as one, in the document *and* in the filename,
    because a PDF of an unpublished roster looks exactly as official as the real
    thing once it has been printed.
"""
from __future__ import annotations

import io
import zipfile

import pytest
from openpyxl import load_workbook

from emma_core.services import render

PUBLISHED = {
    "meta": {
        "description": "Hours worked per staff member",
        "period_start": "2026-08-01", "period_end": "2026-08-28",
        "roster_version_id": "v-1", "roster_version_label": "Cycle 8 v2",
        "roster_version_status": "published",
    },
    "columns": [
        {"key": "staff", "label": "Staff"},
        {"key": "rank", "label": "職級"},
        {"key": "hours", "label": "Hours"},
        {"key": "nights", "label": "N更次數"},
        {"key": "audited", "label": "Medication audited"},
    ],
    "rows": [
        {"staff": "陳小明", "rank": "保健員（HW）", "hours": 176.5,
         "nights": 4, "audited": True},
        {"staff": "Wong Tai Man", "rank": "註冊護士（RN）", "hours": 168,
         "nights": 0, "audited": False},
        {"staff": "李美儀", "rank": "個人照顧員（PCW）", "hours": "180.25",
         "nights": 6, "audited": True},
    ],
}

DRAFT = {**PUBLISHED,
         "meta": {**PUBLISHED["meta"], "roster_version_status": "draft"}}

NO_ROSTER = {
    "meta": {"description": "SWD staff register", "period_start": "2026-08-01",
             "period_end": "2026-08-28"},
    "columns": [{"key": "staff", "label": "Staff"}],
    "rows": [{"staff": "陳小明"}],
}


# ── draft detection ──────────────────────────────────────────────────────────
@pytest.mark.parametrize(("status", "draft"), [
    ("published", False),
    ("draft", True),
    ("archived", True),   # not published is not operative
    ("", False),          # no roster behind the report at all
])
def test_only_a_published_roster_is_not_a_draft(status, draft):
    payload = {**PUBLISHED,
               "meta": {**PUBLISHED["meta"], "roster_version_status": status}}
    assert render.is_draft(payload) is draft


def test_a_report_with_no_roster_is_not_flagged_as_draft():
    """A staff register has no roster version. Watermarking it DRAFT would train
    people to ignore the watermark on the reports where it matters."""
    assert render.is_draft(NO_ROSTER) is False


# ── XLSX ─────────────────────────────────────────────────────────────────────
def _sheet(payload, **kw):
    data = render.to_xlsx(payload, **kw)
    return data, load_workbook(io.BytesIO(data)).active


def test_xlsx_is_a_real_workbook():
    data, sheet = _sheet(PUBLISHED, title="Hours Report")
    assert data[:2] == b"PK", "xlsx must be a zip container"
    assert zipfile.is_zipfile(io.BytesIO(data))
    assert sheet.title == "Hours Report"


def test_numbers_are_written_as_numbers_not_text():
    """The whole reason for shipping xlsx rather than a renamed csv."""
    _, sheet = _sheet(PUBLISHED)
    header = _header_row(sheet)
    hours = _column_values(sheet, header, "Hours")
    assert hours == [176.5, 168, 180.25]
    assert all(isinstance(v, (int, float)) for v in hours)


def test_a_numeric_string_is_coerced():
    """Generators are not consistent about types - 180.25 arrives as a string from
    one query and a float from another. The spreadsheet must not care."""
    _, sheet = _sheet(PUBLISHED)
    header = _header_row(sheet)
    assert _column_values(sheet, header, "Hours")[2] == 180.25


def test_booleans_do_not_become_numbers():
    """True is an int in Python. A Yes/No column silently rendering as 1/0 would
    be summable, sortable and wrong."""
    _, sheet = _sheet(PUBLISHED)
    header = _header_row(sheet)
    assert _column_values(sheet, header, "Medication audited") == ["Yes", "No", "Yes"]


def test_chinese_labels_and_values_survive():
    _, sheet = _sheet(PUBLISHED)
    header = _header_row(sheet)
    assert "職級" in [c.value for c in sheet[header]]
    assert _column_values(sheet, header, "職級")[0] == "保健員（HW）"


def test_the_header_is_frozen_and_filterable():
    """Both matter on a 200-row hours report and neither is decoration."""
    _, sheet = _sheet(PUBLISHED)
    header = _header_row(sheet)
    assert sheet.freeze_panes == f"A{header + 1}"
    assert sheet.auto_filter.ref is not None


def test_the_meta_block_records_what_this_is():
    """A spreadsheet found in a folder a year later has to say which period and
    which roster version it came from, or it is not evidence of anything."""
    _, sheet = _sheet(PUBLISHED, title="Hours Report")
    labels = {sheet.cell(row=r, column=1).value for r in range(1, 12)}
    assert {"Period", "Roster version", "Generated", "Rows"} <= labels


def test_a_draft_workbook_says_so():
    _, sheet = _sheet(DRAFT)
    text = " ".join(str(sheet.cell(row=r, column=c).value or "")
                    for r in range(1, 14) for c in range(1, 3))
    assert "DRAFT" in text
    assert "not been published" in text


def test_an_empty_report_still_produces_a_valid_file():
    """A period with no roster is a normal state, not an error. Returning a
    corrupt file, or a 500, sends people to look for a bug that is not there."""
    empty = {"meta": {"description": "none"}, "columns": [], "rows": []}
    data = render.to_xlsx(empty, title="Empty")
    assert zipfile.is_zipfile(io.BytesIO(data))


def _header_row(sheet) -> int:
    for row in range(1, 20):
        if sheet.cell(row=row, column=1).value == "Staff":
            return row
    raise AssertionError("header row not found")


def _column_values(sheet, header_row, label):
    index = next(c.column for c in sheet[header_row] if c.value == label)
    return [sheet.cell(row=r, column=index).value
            for r in range(header_row + 1, header_row + 4)]


# ── PDF ──────────────────────────────────────────────────────────────────────
def test_pdf_is_a_real_pdf():
    data = render.to_pdf(PUBLISHED, title="Hours Report")
    assert data[:5] == b"%PDF-", "must carry the PDF magic bytes"
    assert b"%%EOF" in data[-1024:]


def test_the_cjk_font_is_embedded():
    """Without MSung-Light every Chinese label renders as a black box. The font is
    built into reportlab, so this asserts it was actually registered and used
    rather than silently falling back to Helvetica."""
    data = render.to_pdf(PUBLISHED, title="Hours Report")
    assert b"MSung-Light" in data


def test_the_traditional_chinese_cmap_is_used_not_the_simplified_one():
    """reportlab's own default pairs MSung-Light with UniGB-UCS2-H - a font in the
    Adobe-CNS1 (Traditional) collection with the Adobe-GB1 (Simplified) CMap. The
    CIDs that produces index into the wrong character collection.

    This is the worst kind of encoding bug: nothing raises, the PDF opens, the
    layout is right, and only a Chinese reader notices the characters are wrong.
    So it gets a test of its own, and one that fails loudly if a reportlab upgrade
    resets the mapping.
    """
    data = render.to_pdf(PUBLISHED, title="Hours Report")
    assert b"UniCNS" in data, "must use the Adobe-CNS1 Traditional Chinese CMap"
    assert b"UniGB" not in data, (
        "UniGB-UCS2-H is the Simplified Chinese CMap and is wrong for MSung-Light")


def test_a_draft_pdf_is_watermarked():
    published = render.to_pdf(PUBLISHED, title="Hours")
    draft = render.to_pdf(DRAFT, title="Hours")
    assert len(draft) > len(published), (
        "the draft should carry extra content - the watermark and the banner")


def test_a_wide_report_goes_landscape():
    """A roster export is staff by days. In portrait it is unreadable, and an
    unreadable PDF gets asked for again as a spreadsheet."""
    wide = {
        "meta": {"description": "Roster export"},
        "columns": [{"key": f"d{i}", "label": f"Day {i}"} for i in range(20)],
        "rows": [{f"d{i}": "A7" for i in range(20)}],
    }
    narrow = {**wide, "columns": wide["columns"][:4],
              "rows": [{f"d{i}": "A7" for i in range(4)}]}
    assert b"/MediaBox [ 0 0 841" in render.to_pdf(wide) or \
           b"841.89" in render.to_pdf(wide), "wide reports should be A4 landscape"
    assert b"595" in render.to_pdf(narrow), "narrow reports should stay portrait"


def test_a_huge_report_is_capped_and_says_so():
    """The cap protects the API process from a 200 000-row layout. Truncating
    quietly would be worse than not rendering at all - the reader has no way to
    know rows are missing, and a short report reads as a clean one."""
    over = 250
    huge = {
        "meta": {"description": "big"},
        "columns": [{"key": "staff", "label": "Staff"}],
        "rows": [{"staff": f"staff-{i}"}
                 for i in range(render.MAX_PDF_ROWS + over)],
    }
    data = render.to_pdf(huge, title="Big")
    assert data[:5] == b"%PDF-"

    # The notice must name both numbers. Extracted rather than searched for in the
    # raw bytes, since the content stream is compressed.
    pdfium = pytest.importorskip(
        "pypdfium2", reason="text extraction needs a PDF reader")
    text = "".join(page.get_textpage().get_text_range()
                   for page in pdfium.PdfDocument(io.BytesIO(data)))
    assert f"{render.MAX_PDF_ROWS:,}" in text
    assert f"{render.MAX_PDF_ROWS + over:,}" in text
    assert "XLSX" in text, "the notice should point at a format that is not capped"


def test_the_uncapped_formats_keep_every_row():
    """The cap is a PDF layout limit, not a data limit. XLSX and CSV must still
    carry the full set, or the notice above is a lie."""
    over = 50
    huge = {
        "meta": {"description": "big"},
        "columns": [{"key": "staff", "label": "Staff"}],
        "rows": [{"staff": f"staff-{i}"}
                 for i in range(render.MAX_PDF_ROWS + over)],
    }
    sheet = load_workbook(io.BytesIO(render.to_xlsx(huge))).active
    assert sheet.max_row >= render.MAX_PDF_ROWS + over


def test_an_empty_report_still_produces_a_valid_pdf():
    empty = {"meta": {"description": "none"}, "columns": [], "rows": []}
    assert render.to_pdf(empty, title="Empty")[:5] == b"%PDF-"


def test_a_report_with_columns_but_no_rows_renders():
    """Distinct from the empty case: a valid report of nothing, which is what a
    zero-violation compliance summary is - and the best possible result."""
    none = {**PUBLISHED, "rows": []}
    assert render.to_pdf(none, title="Compliance")[:5] == b"%PDF-"
    assert zipfile.is_zipfile(io.BytesIO(render.to_xlsx(none, title="Compliance")))


# ── the format registry ──────────────────────────────────────────────────────
def test_all_three_formats_are_registered():
    assert set(render.FORMATS) == {"csv", "xlsx", "pdf"}
    for name, spec in render.FORMATS.items():
        assert spec["media_type"]
        if name != "csv":
            assert callable(spec["render"])


def test_the_media_types_are_the_ones_browsers_and_excel_expect():
    assert render.FORMATS["pdf"]["media_type"] == "application/pdf"
    assert render.FORMATS["xlsx"]["media_type"].endswith("spreadsheetml.sheet")


# ── the download endpoint ────────────────────────────────────────────────────
# Routed as `{report_type}.{fmt}`, so these also pin the path parsing: a dotted
# path segment is easy to get subtly wrong, and the failure looks like a 404 on a
# report that exists.
def _client(monkeypatch, payload=PUBLISHED, title="Hours Report"):
    from fastapi import FastAPI
    from fastapi.testclient import TestClient

    from api.deps import AuthCtx, get_ctx
    from api.routers import reports as reports_router
    from emma_core.models import Profile
    from emma_core.services import reports as reports_svc

    monkeypatch.setattr(
        reports_svc, "generate",
        lambda *a, **k: {"title": title, "payload": payload})

    app = FastAPI()
    app.include_router(reports_router.router)
    app.dependency_overrides[get_ctx] = lambda: AuthCtx(
        token="t", client=object(),
        profile=Profile(id="p1", facility_id="f1", role="superintendent",
                        staff_id=None))
    return TestClient(app)


@pytest.mark.parametrize(("fmt", "media", "magic"), [
    ("csv",  "text/csv",         b"Staff"),
    ("xlsx", "spreadsheetml",    b"PK"),
    ("pdf",  "application/pdf",  b"%PDF-"),
])
def test_every_format_downloads(monkeypatch, fmt, media, magic):
    response = _client(monkeypatch).get(f"/reports/download/roster_hours.{fmt}")
    assert response.status_code == 200
    assert media in response.headers["content-type"]
    assert response.content.startswith(magic) or magic in response.content
    assert f'filename="roster_hours.{fmt}"' in response.headers["content-disposition"]


def test_an_unknown_format_is_a_clean_404(monkeypatch):
    response = _client(monkeypatch).get("/reports/download/roster_hours.docx")
    assert response.status_code == 404
    assert "docx" in response.json()["detail"]["message"]


def test_a_draft_download_is_named_draft(monkeypatch):
    """The marking has to survive leaving the building. A file sitting in someone's
    downloads folder still has to say what it is."""
    response = _client(monkeypatch, payload=DRAFT).get(
        "/reports/download/roster_hours.pdf")
    assert 'filename="roster_hours-DRAFT.pdf"' in response.headers["content-disposition"]


def test_a_published_download_is_not(monkeypatch):
    response = _client(monkeypatch).get("/reports/download/roster_hours.pdf")
    assert "DRAFT" not in response.headers["content-disposition"]


# ── 7.2 · the events overlay on the exported roster (Cherry, 1 Aug 2026) ────
# "A printed roster without special events is incomplete." Two things have to be
# true for the printed sheet to be usable: the event says what it costs, and it
# appears against the floor it was booked for and no other.

class _EventsClient:
    """Answers `facility_events` and `event_staffing_requirements` only."""

    def __init__(self, events, requirements):
        self.data = {"facility_events": events,
                     "event_staffing_requirements": requirements}
        self.name = None

    def table(self, name):
        self.name = name
        return self

    def select(self, *_a, **_k):
        return self

    def eq(self, *_a, **_k):
        return self

    def gte(self, *_a, **_k):
        return self

    def lte(self, *_a, **_k):
        return self

    def in_(self, *_a, **_k):
        return self

    def execute(self):
        return type("R", (), {"data": list(self.data.get(self.name, []))})()


PERIOD = {"period_start": "2026-08-01", "period_end": "2026-08-28"}


def test_an_overlay_states_the_staff_the_event_adds():
    from emma_core.services.reports import _event_overlays

    client = _EventsClient(
        events=[{"id": "e1", "date": "2026-08-04", "title": None,
                 "event_type": "haircut", "unit_id": None}],
        requirements=[{"event_id": "e1", "rank": "CW|HCA", "count": 1,
                       "is_additive": True}],
    )
    overlays = _event_overlays(client, "f1", PERIOD)
    label = overlays["2026-08-04"][0]["label"]
    # The alias the importer writes resolves to the canonical type's label.
    assert "剪髮" in label
    assert "+1 CW|HCA" in label, "the printed roster must say what the event costs"


def test_a_concurrent_duty_is_not_advertised_as_an_extra_body():
    """Podiatry is covered by staff already rostered. Printing '+1 HW' beside it
    would have the manager rostering a head the rule never asked for."""
    from emma_core.services.reports import _event_overlays

    client = _EventsClient(
        events=[{"id": "e1", "date": "2026-08-04", "title": "足部護理",
                 "event_type": "podiatry", "unit_id": None}],
        requirements=[{"event_id": "e1", "rank": "HW", "count": 1,
                       "is_additive": False}],
    )
    assert _event_overlays(client, "f1", PERIOD)["2026-08-04"][0]["label"] == "足部護理"


def test_an_overlay_remembers_which_floor_booked_it():
    from emma_core.services.reports import _event_overlays

    client = _EventsClient(
        events=[{"id": "e1", "date": "2026-08-04", "title": "CGAT",
                 "event_type": "cgat", "unit_id": "unit-3f"}],
        requirements=[],
    )
    assert _event_overlays(client, "f1", PERIOD)["2026-08-04"][0]["unit_id"] == "unit-3f"


def test_missing_requirements_degrade_to_a_plain_title():
    """Losing the overlay detail is a defect; losing the roster is an outage."""
    from emma_core.services.reports import _event_overlays

    class _NoRequirements(_EventsClient):
        def execute(self):
            if self.name == "event_staffing_requirements":
                raise RuntimeError("table is unreachable")
            return super().execute()

    client = _NoRequirements(
        events=[{"id": "e1", "date": "2026-08-04", "title": "剪髮",
                 "event_type": "hair_cutting", "unit_id": None}],
        requirements=[],
    )
    assert _event_overlays(client, "f1", PERIOD)["2026-08-04"][0]["label"] == "剪髮"
